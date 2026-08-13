"""Builds the downloadable result workbook.

Input-sheet mode:
    Bin Recommendations - every Input row plus Recommended Bin, Status,
                          utilisation and the recommendation reason
    Bin Rule Master     - the six categories and how many products landed in each
    Summary             - headline metrics for the run
    Validation Issues   - every warning and error raised during processing

Master-bin mode:
    Part Requirements   - the original columns plus Bin Suggestion / Suggestion Reason
    Master              - the bin master, with how many parts landed in each bin
    Summary             - headline metrics for the run
    Validation Issues   - every warning and error raised during processing
    Logic Guide         - carried through from the uploaded workbook when present

The uploaded workbook is never written to; this always builds a fresh file.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
from core.models import (
    MODE_INPUT,
    STATUS_ASSIGNED,
    STATUS_ERROR,
    STATUS_INVALID_DATA,
    STATUS_MATCHED,
    STATUS_UNASSIGNED,
    AnalysisResult,
)

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F3A5F")

ASSIGNED_FILL = PatternFill("solid", fgColor="E7F5EC")
UNASSIGNED_FILL = PatternFill("solid", fgColor="FDECEC")
ERROR_FILL = PatternFill("solid", fgColor="FFF4E0")

THIN = Side(style="thin", color="D6DCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

INT_FORMAT = "#,##0"
WEIGHT_FORMAT = "#,##0.000"
PERCENT_FORMAT = "0.0%"

PART_HEADERS = [
    ("Part Number", 16, None),
    ("Part Description", 26, None),
    (f"Length ({config.DIMENSION_UNIT})", 12, INT_FORMAT),
    (f"Breadth ({config.DIMENSION_UNIT})", 13, INT_FORMAT),
    (f"Width ({config.DIMENSION_UNIT})", 12, INT_FORMAT),
    (f"Total Volume ({config.VOLUME_UNIT})", 18, INT_FORMAT),
    ("ROP (Qty)", 11, INT_FORMAT),
    (f"Required Volume ({config.VOLUME_UNIT})", 20, INT_FORMAT),
    (f"Material Weight/Unit ({config.WEIGHT_UNIT})", 22, WEIGHT_FORMAT),
    (f"Required Weight ({config.WEIGHT_UNIT})", 19, WEIGHT_FORMAT),
    ("Bin Suggestion", 16, None),
    ("Suggestion Reason", 80, None),
    ("Status", 13, None),
    ("Bin Location", 14, None),
    ("Volume Utilisation", 17, PERCENT_FORMAT),
    ("Weight Utilisation", 17, PERCENT_FORMAT),
    ("Orientation Applied", 18, None),
    ("Alternative Bins", 22, None),
]

BIN_HEADERS = [
    ("Bin ID", 14, None),
    ("Bin Description", 22, None),
    (f"Bin Length ({config.DIMENSION_UNIT})", 16, INT_FORMAT),
    (f"Bin Breadth ({config.DIMENSION_UNIT})", 17, INT_FORMAT),
    (f"Bin Width ({config.DIMENSION_UNIT})", 16, INT_FORMAT),
    (f"Cubic Capacity ({config.VOLUME_UNIT})", 20, INT_FORMAT),
    (f"Max Weight ({config.WEIGHT_UNIT})", 16, WEIGHT_FORMAT),
    ("Location", 14, None),
    ("Status", 13, None),
    ("Parts Assigned", 15, INT_FORMAT),
]

ISSUE_HEADERS = [
    ("Severity", 12, None),
    ("Sheet", 20, None),
    ("Row", 8, INT_FORMAT),
    ("Reference", 16, None),
    ("Message", 100, None),
]

DECIMAL_FORMAT = "#,##0.00"

# Input-sheet mode.  The source columns are reproduced first so the sheet reads
# like the uploaded Input sheet, then the seven engine outputs are appended.
INPUT_HEADERS = [
    ("S No", 8, None),
    ("DWG No", 18, None),
    ("SAP No", 14, None),
    ("Description", 34, None),
    ("Qty/Mc", 9, DECIMAL_FORMAT),
    ("ROP", 10, DECIMAL_FORMAT),
    (f"Length ({config.DIMENSION_UNIT})", 12, DECIMAL_FORMAT),
    (f"Breadth ({config.DIMENSION_UNIT})", 12, DECIMAL_FORMAT),
    (f"Height ({config.DIMENSION_UNIT})", 12, DECIMAL_FORMAT),
    ("Weight (Kg)", 12, DECIMAL_FORMAT),
    ("Product Volume (mm^3)", 20, INT_FORMAT),
    ("Demand Product Volume (mm^3)", 24, INT_FORMAT),
    ("Demand Product Weight (Kg)", 22, DECIMAL_FORMAT),
    ("Recommended Bin", 18, None),
    ("Recommendation Status", 19, None),
    ("Volume Utilization %", 18, DECIMAL_FORMAT),
    ("Weight Utilization %", 18, DECIMAL_FORMAT),
    ("Recommendation Reason", 100, None),
]

RULE_HEADERS = [
    ("Bin Type", 20, None),
    ("Minimum Volume (mm^3)", 20, INT_FORMAT),
    ("Maximum Volume (mm^3)", 20, INT_FORMAT),
    ("Minimum Weight (Kg)", 18, DECIMAL_FORMAT),
    ("Maximum Weight (Kg)", 18, DECIMAL_FORMAT),
    ("Priority", 10, INT_FORMAT),
    ("Products Assigned", 17, INT_FORMAT),
]


def _write_header(worksheet, headers, row: int = 1) -> None:
    for index, (title, width, _) in enumerate(headers, start=1):
        cell = worksheet.cell(row=row, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.row_dimensions[row].height = 30
    worksheet.freeze_panes = worksheet.cell(row=row + 1, column=1)


def _write_row(worksheet, row_index: int, headers, values, fill=None) -> None:
    for index, value in enumerate(values, start=1):
        cell = worksheet.cell(row=row_index, column=index, value=value)
        number_format = headers[index - 1][2]
        if number_format and isinstance(value, (int, float)):
            cell.number_format = number_format
        cell.alignment = Alignment(vertical="top", wrap_text=(headers[index - 1][1] >= 40))
        cell.border = BORDER
        if fill is not None:
            cell.fill = fill


def _sheet_parts(workbook: Workbook, result: AnalysisResult) -> None:
    worksheet = workbook.create_sheet("Part Requirements")
    _write_header(worksheet, PART_HEADERS)

    fills = {
        STATUS_ASSIGNED: ASSIGNED_FILL,
        STATUS_UNASSIGNED: UNASSIGNED_FILL,
        STATUS_ERROR: ERROR_FILL,
    }

    for offset, rec in enumerate(result.recommendations, start=2):
        part = rec.part
        _write_row(worksheet, offset, PART_HEADERS, [
            part.part_number,
            part.description,
            part.length,
            part.breadth,
            part.width,
            part.unit_volume,
            part.rop,
            part.required_volume,
            part.weight_per_unit,
            part.required_weight,
            rec.bin_suggestion,
            rec.reason,
            rec.status.title(),
            rec.recommended_bin.location if rec.recommended_bin else "",
            rec.volume_utilisation,
            rec.weight_utilisation,
            "Yes" if rec.orientation_used else ("No" if rec.status == STATUS_ASSIGNED else ""),
            ", ".join(rec.alternatives),
        ], fill=fills.get(rec.status))

    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(PART_HEADERS))}{len(result.recommendations) + 1}"


def _sheet_bins(workbook: Workbook, result: AnalysisResult) -> None:
    worksheet = workbook.create_sheet("Master")
    _write_header(worksheet, BIN_HEADERS)

    load = result.summary.get("bin_load", {})
    for offset, storage_bin in enumerate(result.bins, start=2):
        _write_row(worksheet, offset, BIN_HEADERS, [
            storage_bin.bin_id,
            storage_bin.description,
            storage_bin.length,
            storage_bin.breadth,
            storage_bin.width,
            storage_bin.usable_capacity,
            storage_bin.max_weight,
            storage_bin.location,
            storage_bin.status,
            load.get(storage_bin.bin_id, 0),
        ], fill=None if storage_bin.is_available else ERROR_FILL)


def _sheet_summary(workbook: Workbook, result: AnalysisResult) -> None:
    worksheet = workbook.create_sheet("Summary", 0)
    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 30

    title = worksheet.cell(row=1, column=1, value="Material Bin Recommendation - Summary")
    title.font = TITLE_FONT

    summary = result.summary
    average = summary.get("average_volume_utilisation")

    rows = [
        ("Source workbook", result.source_name or "uploaded workbook"),
        ("Generated on", datetime.now().strftime("%d-%b-%Y %H:%M")),
        ("", ""),
        ("Total parts analysed", summary["total_parts"]),
        (f"Total required volume ({config.VOLUME_UNIT})", summary["total_required_volume"]),
        (f"Total required weight ({config.WEIGHT_UNIT})", summary["total_required_weight"]),
        ("Successfully assigned bins", summary["assigned"]),
        ("Unassigned materials", summary["unassigned"]),
        ("  - no suitable bin", summary["no_suitable_bin"]),
        ("  - data errors", summary["data_errors"]),
        ("", ""),
        ("Bins in master", summary["total_bins"]),
        ("Bins available for matching", summary["available_bins"]),
        ("Distinct bins used", summary["distinct_bins_used"]),
        ("Average volume utilisation", average),
        ("Assignments needing rotation", summary["orientation_used_count"]),
        ("Validation issues raised", summary["issue_count"]),
    ]

    for offset, (label, value) in enumerate(rows, start=3):
        label_cell = worksheet.cell(row=offset, column=1, value=label)
        label_cell.font = Font(bold=not label.startswith("  ") and bool(label))
        value_cell = worksheet.cell(row=offset, column=2, value=value)
        if isinstance(value, float) and label.startswith("Average"):
            value_cell.number_format = PERCENT_FORMAT
        elif isinstance(value, (int, float)):
            value_cell.number_format = WEIGHT_FORMAT if "weight" in label.lower() else INT_FORMAT


def _sheet_issues(workbook: Workbook, result: AnalysisResult) -> None:
    if not result.issues:
        return
    worksheet = workbook.create_sheet("Validation Issues")
    _write_header(worksheet, ISSUE_HEADERS)
    for offset, issue in enumerate(result.issues, start=2):
        _write_row(worksheet, offset, ISSUE_HEADERS, [
            issue.severity.title(),
            issue.sheet,
            issue.row,
            issue.reference,
            issue.message,
        ], fill=ERROR_FILL if issue.severity == "error" else None)


def _sheet_guide(workbook: Workbook, result: AnalysisResult) -> None:
    if not result.guide_rows:
        return
    worksheet = workbook.create_sheet("Logic Guide")
    widths = [24, 40, 70]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    for row_index, values in enumerate(result.guide_rows, start=1):
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
    worksheet.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Input-sheet mode
# ---------------------------------------------------------------------------

def _sheet_recommendations(workbook: Workbook, result: AnalysisResult) -> None:
    worksheet = workbook.create_sheet("Bin Recommendations")
    _write_header(worksheet, INPUT_HEADERS)

    fills = {
        STATUS_MATCHED: ASSIGNED_FILL,
        STATUS_INVALID_DATA: ERROR_FILL,
    }

    for offset, item in enumerate(result.materials, start=2):
        m, d = item.material, item.decision
        _write_row(worksheet, offset, INPUT_HEADERS, [
            m.s_no,
            m.dwg_no,
            m.sap_no,
            m.description,
            m.qty_per_mc,
            m.rop,
            m.length,
            m.breadth,
            m.height,
            m.weight,
            m.product_volume,
            d.demand_volume if d.demand_volume is not None else m.demand_volume,
            d.demand_weight if d.demand_weight is not None else m.demand_weight,
            d.recommended_bin,
            d.status,
            None if d.volume_utilisation_pct is None else round(d.volume_utilisation_pct, 2),
            None if d.weight_utilisation_pct is None else round(d.weight_utilisation_pct, 2),
            d.reason,
        ], fill=fills.get(d.status, UNASSIGNED_FILL))

    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(INPUT_HEADERS))}{len(result.materials) + 1}"


def _sheet_rules(workbook: Workbook, result: AnalysisResult) -> None:
    worksheet = workbook.create_sheet("Bin Rule Master")
    _write_header(worksheet, RULE_HEADERS)

    load = result.summary.get("bin_load", {})
    for offset, rule in enumerate(result.rules, start=2):
        _write_row(worksheet, offset, RULE_HEADERS, [
            rule.name,
            rule.min_volume,
            rule.max_volume,
            rule.min_weight,
            rule.max_weight,
            rule.priority + 1,
            load.get(rule.name, 0),
        ])

    note = worksheet.cell(
        row=len(result.rules) + 3, column=1,
        value="A category is eligible only when the demand volume AND the demand weight are both "
              "inside its range. Bounds are inclusive; Priority breaks ties where two bands meet.",
    )
    note.font = Font(italic=True, size=10)


def _sheet_input_summary(workbook: Workbook, result: AnalysisResult) -> None:
    worksheet = workbook.create_sheet("Summary", 0)
    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 30

    title = worksheet.cell(row=1, column=1, value="Material Bin Recommendation - Summary")
    title.font = TITLE_FONT

    s = result.summary
    rows = [
        ("Source workbook", result.source_name or "uploaded workbook"),
        ("Generated on", datetime.now().strftime("%d-%b-%Y %H:%M")),
        ("", ""),
        ("Total products", s["total_products"]),
        ("Matched products", s["matched_products"]),
        ("No suitable bin", s["no_suitable_bin"]),
        ("Invalid data rows", s["invalid_products"]),
        ("", ""),
        ("Total demand volume (mm^3)", s["total_demand_volume"]),
        ("Total demand weight (Kg)", s["total_demand_weight"]),
        ("Average volume utilisation %", s["average_volume_utilisation_pct"]),
        ("Average weight utilisation %", s["average_weight_utilisation_pct"]),
        ("Validation issues raised", s["issue_count"]),
        ("", ""),
        ("Products per bin type", ""),
    ]
    rows.extend((f"  - {name}", count) for name, count in s.get("bin_load", {}).items())

    for offset, (label, value) in enumerate(rows, start=3):
        label_cell = worksheet.cell(row=offset, column=1, value=label)
        label_cell.font = Font(bold=not label.startswith("  ") and bool(label))
        value_cell = worksheet.cell(row=offset, column=2, value=value)
        if isinstance(value, (int, float)):
            value_cell.number_format = (
                DECIMAL_FORMAT if ("weight" in label.lower() or "%" in label) else INT_FORMAT
            )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def build_workbook(result: AnalysisResult) -> bytes:
    """Render an AnalysisResult as .xlsx bytes."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    if result.mode == MODE_INPUT:
        _sheet_recommendations(workbook, result)
        _sheet_rules(workbook, result)
        _sheet_issues(workbook, result)
        _sheet_guide(workbook, result)
        _sheet_input_summary(workbook, result)   # inserted at position 0
    else:
        _sheet_parts(workbook, result)
        _sheet_bins(workbook, result)
        _sheet_issues(workbook, result)
        _sheet_guide(workbook, result)
        _sheet_summary(workbook, result)         # inserted at position 0

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def output_filename(source_name: str = "") -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = "Bin_Recommendations"
    if source_name:
        base = source_name.rsplit(".", 1)[0]
        base = "".join(ch for ch in base if ch.isalnum() or ch in " _-").strip()
        if base:
            stem = f"{base}_Recommendations"
    return f"{stem}_{stamp}.xlsx"

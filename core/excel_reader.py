"""Workbook parsing.

Turns an uploaded .xlsx into InputMaterial / PartRequirement / StorageBin
objects.  Nothing in here knows about matching rules - it only locates the
sheets, maps the columns and coerces the cell values, recording a
ValidationIssue whenever something looks wrong.

Two workbook layouts are recognised, and the sheets present decide which:

    Input             a flat sheet of materials that already carries the demand
                      volume and demand weight, matched against the bin rule
                      master in the persistent rule store (seeded from
                      config.BIN_RULES)
    Part Requirements + Master
                      materials sized from ROP and matched against a bin list
                      supplied in the workbook

Sheets are read with openpyxl in read-only mode so a workbook with tens of
thousands of rows streams instead of being materialised twice.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

import config
from core.models import (
    MODE_INPUT,
    MODE_MASTER,
    InputMaterial,
    PartRequirement,
    StorageBin,
    ValidationIssue,
)

# How many leading rows to search when looking for the header row.
HEADER_SEARCH_DEPTH = 15

_PAREN = re.compile(r"\([^)]*\)")
_NON_ALNUM = re.compile(r"[^a-z0-9]+")


class WorkbookError(Exception):
    """Raised when the workbook cannot be used at all."""


@dataclass
class WorkbookContents:
    """Everything read off one workbook, plus which layout it turned out to be."""

    mode: str
    issues: List[ValidationIssue] = field(default_factory=list)
    guide_rows: List[List[str]] = field(default_factory=list)

    # MODE_INPUT
    materials: List[InputMaterial] = field(default_factory=list)

    # MODE_MASTER
    parts: List[PartRequirement] = field(default_factory=list)
    bins: List[StorageBin] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------

def normalise(value) -> str:
    """Fold a header or sheet name to a comparable key.

    "Length (mm)"              -> "length"
    "Material Weight/Unit (kg)" -> "materialweightunit"
    """
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = _PAREN.sub(" ", text)
    text = text.replace("³", "3").replace("²", "2")
    return _NON_ALNUM.sub("", text)


def to_number(value) -> Optional[float]:
    """Coerce a cell to a float, tolerating "1,200" or "300 mm"."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_blank_row(values: Sequence) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


# ---------------------------------------------------------------------------
# Sheet + column resolution
# ---------------------------------------------------------------------------

def resolve_sheet(sheet_names: Sequence[str], aliases: Sequence[str]) -> Optional[str]:
    """Pick the sheet best matching a list of aliases.

    Exact normalised matches win; a partial (substring) match is the fallback
    so "Part Requirements FY26" still resolves.
    """
    normalised = {name: normalise(name) for name in sheet_names}

    for alias in aliases:
        for name, key in normalised.items():
            if key == alias:
                return name

    for alias in aliases:
        for name, key in normalised.items():
            if alias in key or key in alias:
                return name
    return None


def _score_header_row(values: Sequence, alias_map: Dict[str, Sequence[str]]) -> int:
    """How many known fields a candidate header row resolves."""
    keys = {normalise(v) for v in values if v is not None}
    keys.discard("")
    return sum(1 for aliases in alias_map.values() if keys.intersection(aliases))


def map_columns(header: Sequence, alias_map: Dict[str, Sequence[str]]) -> Dict[str, int]:
    """Map internal field names to column indexes in the header row.

    Fields are resolved in declaration order and a column is only claimed once,
    so overlapping spellings (Breadth vs Width) settle deterministically.
    """
    keys = [normalise(v) for v in header]
    claimed: set = set()
    mapping: Dict[str, int] = {}

    for field, aliases in alias_map.items():
        for alias in aliases:
            for index, key in enumerate(keys):
                if index in claimed or not key:
                    continue
                if key == alias:
                    mapping[field] = index
                    claimed.add(index)
                    break
            if field in mapping:
                break
    return mapping


def _read_sheet(worksheet, alias_map: Dict[str, Sequence[str]]) -> Tuple[Dict[str, int], int, Iterable]:
    """Locate the header row and return (column map, header row number, row iterator)."""
    rows = worksheet.iter_rows(values_only=True)
    buffered: List[Tuple[int, Sequence]] = []

    best_index = -1
    best_row: Sequence = ()
    best_score = 0

    for offset, values in enumerate(rows):
        buffered.append((offset + 1, values))
        if _is_blank_row(values):
            if offset + 1 >= HEADER_SEARCH_DEPTH:
                break
            continue
        score = _score_header_row(values, alias_map)
        if score > best_score:
            best_score, best_index, best_row = score, offset, values
        if offset + 1 >= HEADER_SEARCH_DEPTH:
            break

    if best_score < 2:
        raise WorkbookError(
            f"Could not find a recognisable header row on sheet '{worksheet.title}'."
        )

    mapping = map_columns(best_row, alias_map)
    header_row_number = buffered[best_index][0]

    def iter_data():
        # Rows already pulled from the stream while hunting for the header.
        for row_number, values in buffered[best_index + 1:]:
            yield row_number, values
        offset = len(buffered)
        for values in rows:
            offset += 1
            yield offset, values

    return mapping, header_row_number, iter_data()


def _cell(values: Sequence, mapping: Dict[str, int], field: str):
    index = mapping.get(field)
    if index is None or index >= len(values):
        return None
    return values[index]


# ---------------------------------------------------------------------------
# Input sheet (bin rule master mode)
# ---------------------------------------------------------------------------

def read_input_materials(worksheet, issues: List[ValidationIssue]) -> List[InputMaterial]:
    """Read every row of the Input sheet.

    Values are taken as they stand - nothing is recalculated here, because the
    demand volume and demand weight in the sheet are the primary inputs to the
    recommendation.  Rows stream one at a time, so a 5,000 row sheet costs no
    more memory per row than a 100 row one.
    """
    mapping, header_row, rows = _read_sheet(worksheet, config.INPUT_COLUMN_ALIASES)
    sheet = worksheet.title.strip() or worksheet.title

    missing = [f for f in config.REQUIRED_INPUT_FIELDS if f not in mapping]
    if missing:
        labels = {
            "demand_volume": "Demand product volume (mm^3)",
            "demand_weight": "demand product weights (Kg)",
        }
        raise WorkbookError(
            f"Sheet '{sheet}' is missing required column(s): "
            + ", ".join(labels.get(f, f) for f in missing)
            + ". These two columns are the primary inputs for the bin recommendation."
        )

    for optional in ("s_no", "dwg_no", "sap_no", "description", "rop"):
        if optional not in mapping:
            issues.append(ValidationIssue(
                sheet, header_row, "", "warning",
                f"Column '{optional}' was not found; it will be blank in the results.",
            ))

    materials: List[InputMaterial] = []

    for row_number, values in rows:
        if _is_blank_row(values):
            continue

        raw_volume = _cell(values, mapping, "demand_volume")
        raw_weight = _cell(values, mapping, "demand_weight")

        material = InputMaterial(
            row=row_number,
            s_no=to_text(_cell(values, mapping, "s_no")),
            dwg_no=to_text(_cell(values, mapping, "dwg_no")),
            sap_no=to_text(_cell(values, mapping, "sap_no")),
            description=to_text(_cell(values, mapping, "description")),
            qty_per_mc=to_number(_cell(values, mapping, "qty_per_mc")),
            rop=to_number(_cell(values, mapping, "rop")),
            length=to_number(_cell(values, mapping, "length")),
            breadth=to_number(_cell(values, mapping, "breadth")),
            height=to_number(_cell(values, mapping, "height")),
            weight=to_number(_cell(values, mapping, "weight")),
            product_volume=to_number(_cell(values, mapping, "product_volume")),
            demand_volume=to_number(raw_volume),
            demand_weight=to_number(raw_weight),
            raw_demand_volume=raw_volume,
            raw_demand_weight=raw_weight,
        )
        materials.append(material)

    if not materials:
        raise WorkbookError(f"Sheet '{sheet}' contains no material rows.")
    return materials


# ---------------------------------------------------------------------------
# Part Requirements
# ---------------------------------------------------------------------------

def read_parts(worksheet, issues: List[ValidationIssue]) -> List[PartRequirement]:
    mapping, _, rows = _read_sheet(worksheet, config.PART_COLUMN_ALIASES)
    sheet = worksheet.title

    missing_required = [f for f in config.REQUIRED_PART_FIELDS if f not in mapping]
    if missing_required:
        raise WorkbookError(
            f"Sheet '{sheet}' is missing required column(s): {', '.join(missing_required)}."
        )

    for field in ("length", "breadth", "width", "weight_per_unit"):
        if field not in mapping:
            issues.append(ValidationIssue(
                sheet, None, "", "warning",
                f"Column '{field}' was not found; checks that depend on it will be skipped.",
            ))

    parts: List[PartRequirement] = []
    seen: Dict[str, int] = {}

    for row_number, values in rows:
        if _is_blank_row(values):
            continue

        part_number = to_text(_cell(values, mapping, "part_number"))
        if not part_number:
            issues.append(ValidationIssue(
                sheet, row_number, "", "warning",
                "Row skipped because it has no Part Number.",
            ))
            continue

        part = PartRequirement(
            row=row_number,
            part_number=part_number,
            description=to_text(_cell(values, mapping, "description")),
            length=to_number(_cell(values, mapping, "length")),
            breadth=to_number(_cell(values, mapping, "breadth")),
            width=to_number(_cell(values, mapping, "width")),
            rop=to_number(_cell(values, mapping, "rop")),
            weight_per_unit=to_number(_cell(values, mapping, "weight_per_unit")),
            source_unit_volume=to_number(_cell(values, mapping, "unit_volume")),
            source_required_volume=to_number(_cell(values, mapping, "required_volume")),
            source_required_weight=to_number(_cell(values, mapping, "required_weight")),
        )

        if part_number in seen:
            issues.append(ValidationIssue(
                sheet, row_number, part_number, "warning",
                f"Duplicate Part Number (also on row {seen[part_number]}); both rows are analysed.",
            ))
        else:
            seen[part_number] = row_number

        parts.append(part)

    if not parts:
        raise WorkbookError(f"Sheet '{sheet}' contains no material rows.")
    return parts


# ---------------------------------------------------------------------------
# Master bins
# ---------------------------------------------------------------------------

def read_bins(worksheet, issues: List[ValidationIssue]) -> List[StorageBin]:
    mapping, _, rows = _read_sheet(worksheet, config.BIN_COLUMN_ALIASES)
    sheet = worksheet.title

    if "bin_id" not in mapping:
        raise WorkbookError(f"Sheet '{sheet}' is missing a Bin ID column.")
    if "cubic_capacity" not in mapping and not {"length", "breadth", "width"}.issubset(mapping):
        raise WorkbookError(
            f"Sheet '{sheet}' needs either a Cubic Capacity column or all three bin dimensions."
        )

    bins: List[StorageBin] = []
    seen: Dict[str, int] = {}

    for row_number, values in rows:
        if _is_blank_row(values):
            continue

        bin_id = to_text(_cell(values, mapping, "bin_id"))
        if not bin_id:
            issues.append(ValidationIssue(
                sheet, row_number, "", "warning",
                "Row skipped because it has no Bin ID.",
            ))
            continue

        if bin_id in seen:
            issues.append(ValidationIssue(
                sheet, row_number, bin_id, "warning",
                f"Duplicate Bin ID (also on row {seen[bin_id]}); this row is ignored.",
            ))
            continue
        seen[bin_id] = row_number

        status_text = to_text(_cell(values, mapping, "status"))
        status_key = normalise(status_text)
        if not status_key:
            is_available = config.TREAT_BLANK_STATUS_AS_AVAILABLE
            if not is_available:
                issues.append(ValidationIssue(
                    sheet, row_number, bin_id, "warning",
                    "Bin has a blank Status and is treated as unavailable.",
                ))
        else:
            is_available = status_key in config.AVAILABLE_STATUS_VALUES

        bins.append(StorageBin(
            row=row_number,
            bin_id=bin_id,
            description=to_text(_cell(values, mapping, "description")),
            length=to_number(_cell(values, mapping, "length")),
            breadth=to_number(_cell(values, mapping, "breadth")),
            width=to_number(_cell(values, mapping, "width")),
            cubic_capacity=to_number(_cell(values, mapping, "cubic_capacity")),
            max_weight=to_number(_cell(values, mapping, "max_weight")),
            location=to_text(_cell(values, mapping, "location")),
            status=status_text or ("Available" if is_available else "Unknown"),
            is_available=is_available,
        ))

    if not bins:
        raise WorkbookError(f"Sheet '{sheet}' contains no bin rows.")
    return bins


# ---------------------------------------------------------------------------
# Logic Guide (documentation sheet, carried through untouched)
# ---------------------------------------------------------------------------

def read_guide(worksheet, limit: int = 100) -> List[List[str]]:
    rows: List[List[str]] = []
    for values in worksheet.iter_rows(values_only=True):
        if _is_blank_row(values):
            continue
        rows.append([to_text(v) for v in values])
        if len(rows) >= limit:
            break
    return rows


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def _read_input_mode(workbook, input_sheet: str, issues: List[ValidationIssue]) -> WorkbookContents:
    guide_sheet = resolve_sheet(workbook.sheetnames, config.GUIDE_SHEET_ALIASES)
    if guide_sheet == input_sheet:
        guide_sheet = None

    return WorkbookContents(
        mode=MODE_INPUT,
        issues=issues,
        materials=read_input_materials(workbook[input_sheet], issues),
        guide_rows=read_guide(workbook[guide_sheet]) if guide_sheet else [],
    )


def _read_master_mode(workbook, issues: List[ValidationIssue]) -> WorkbookContents:
    names = workbook.sheetnames
    part_sheet = resolve_sheet(names, config.PART_SHEET_ALIASES)
    bin_sheet = resolve_sheet(names, config.BIN_SHEET_ALIASES)
    guide_sheet = resolve_sheet(names, config.GUIDE_SHEET_ALIASES)

    # Never let one sheet play two roles.
    if bin_sheet == part_sheet:
        bin_sheet = None
    if guide_sheet in (part_sheet, bin_sheet):
        guide_sheet = None

    if part_sheet is None:
        raise WorkbookError(
            "No usable sheet found. Expected either an 'Input' sheet carrying the demand "
            "volume and demand weight, or a 'Part Requirements' sheet with a 'Master' bin "
            f"sheet. Sheets present: {', '.join(names)}."
        )
    if bin_sheet is None:
        raise WorkbookError(
            "No Master bin sheet found. Expected a sheet named like 'Master'. "
            f"Sheets present: {', '.join(names)}."
        )

    return WorkbookContents(
        mode=MODE_MASTER,
        issues=issues,
        parts=read_parts(workbook[part_sheet], issues),
        bins=read_bins(workbook[bin_sheet], issues),
        guide_rows=read_guide(workbook[guide_sheet]) if guide_sheet else [],
    )


def read_workbook(stream) -> WorkbookContents:
    """Read an .xlsx stream, choosing the layout from the sheets it contains.

    An 'Input' sheet wins when present, because it carries the demand figures
    the bin rule master needs.  Otherwise the Part Requirements + Master
    layout is used.
    """
    issues: List[ValidationIssue] = []

    try:
        workbook = load_workbook(stream, data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises a variety of types
        raise WorkbookError(f"The file could not be opened as an Excel workbook: {exc}") from exc

    try:
        input_sheet = resolve_sheet(workbook.sheetnames, config.INPUT_SHEET_ALIASES)
        if input_sheet is not None:
            try:
                return _read_input_mode(workbook, input_sheet, issues)
            except WorkbookError:
                # A sheet whose name merely looks like "Input" but carries no
                # demand columns must not shadow a valid Master layout.
                if resolve_sheet(workbook.sheetnames, config.PART_SHEET_ALIASES) is None:
                    raise
                issues.clear()
        return _read_master_mode(workbook, issues)
    finally:
        workbook.close()
